#!/usr/bin/env python3
"""
autoacb.py — Canadian ACB / Capital Gains Calculator (Excel-based)

INPUT:  an .xlsx workbook with two required sheets:
  - "Transactions"  : Date | Symbol | settlementDate | Type | Quantity | Price | Fee | Currency
                       When both "Date" and "settlementDate" exist, settlementDate is
                       used (not the trade date). Fee is optional; missing -> $0.
  - "ExchangeRate"   : pasted straight from the Bank of Canada Valet CSV download
                       (https://www.bankofcanada.ca/valet/observations/FXUSDCAD,FXEURCAD/csv?...) —
                       date | FXUSDCAD | FXEURCAD | ... (any FX<CCY>CAD column is auto-detected)
                       CAD itself needs no rows — it's always rate 1. Lookup uses the
                       most recent rate on or before the transaction date.

  Action/Type is BUY, SELL, ROC, or SPLIT (plus common broker synonyms — DIY_BUY,
  ReturnOfCapital, Consolidation, etc.). Anything else (dividends, transfers, interest...)
  is silently skipped — not included anywhere in the output, just noted in a one-line
  console summary so you can see what got left out.
  - Quantity can be positive or negative (source data often uses negative for a sell) —
    Action/Type is authoritative for direction; the sign is normalized and flagged in Notes.
  - ROC (return of capital, e.g. a T3 slip Box 42): leave Quantity blank (defaults to 1)
    and put the TOTAL amount received in Price. Reduces ACB without changing share count;
    if it would push ACB negative, the excess is a deemed capital gain (CRA s.53(2)(h))
    and ACB resets to $0. If T3 provides a CAD amount but the security is not in CAD, 
    convert it to the security's currency first (e.g. USD) and enter that in Price, because 
    the FX rate is applied to the Price to get the CAD amount. 
  - SPLIT (covers reverse splits / consolidations too): Quantity = new shares,
    Price = old shares — a 5-for-1 split is Quantity=5, Price=1; a 10-for-1 reverse
    split is Quantity=1, Price=10. Share count multiplies by Quantity/Price; ACB/share
    moves inversely; total ACB is unchanged.

OUTPUT: a new .xlsx workbook = your input workbook + new sheets:
  - "Schedule 3 - <year>" : one per year with reportable activity, matching CRA Schedule 3
                            (publicly traded shares section) columns, for direct transcription.
  - "T1135 - <year>"      : one per year with any holding — tickers as rows, months as
                            columns, for the $100,000 specified-foreign-property test.
                            Non-CAD tickers default to "Yes" reportable, CAD tickers
                            default to "No" (since currency alone doesn't decide
                            specified-foreign-property status — e.g. a CAD-hedged CDR
                            of a foreign company is still reportable). Edit the
                            "T1135 Reportable?" column directly in this sheet — the
                            monthly totals are live SUMIFS formulas that update
                            automatically when you change Yes/No. A holding with no
                            recent transactions still carries its ACB forward through
                            the current month (run date) rather than stopping at the
                            last transaction — so the current year's tab stays current
                            even during a quiet month, and a position with no activity
                            at all this year still gets a tab. Future months are left
                            blank rather than assumed unchanged.
  - "Summary"             : full chronological ACB ledger, ticker by ticker — every
                            transaction, running ACB, running share balance, and
                            superficial-loss detail, reviewable line by line.

Usage:
    python3 autoacb.py sourcefile.xlsx

All years present in the data are processed automatically — there's no year argument.
"""
import argparse
import re
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

L = get_column_letter   # short alias, used heavily below for building formulas

FONT_NAME = "Arial"
MONEY_FMT = '$#,##0.00;($#,##0.00)'
MONEY_FMT_4DP = '$#,##0.0000;($#,##0.0000)'
QTY_FMT = '#,##0.####'
RATE_FMT = '0.0000'
DATE_FMT = 'yyyy-mm-dd'
PCT_FMT = '0.0%'
T1135_THRESHOLD = Decimal('100000')
CENT = Decimal('0.01')
UNIT4DP = Decimal('0.0001')


def _r2(d):
    return d.quantize(CENT, rounding=ROUND_HALF_UP)


def _r4(d):
    return d.quantize(UNIT4DP, rounding=ROUND_HALF_UP)


HEADER_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF')
BAND_FILL = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
BAND_FONT = Font(name=FONT_NAME, bold=True, size=12)
SL_FILL = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
WARN_FILL = PatternFill('solid', start_color='F8CBAD', end_color='F8CBAD')
EXCLUDED_FILL = PatternFill('solid', start_color='F2F2F2', end_color='F2F2F2')
SUBTOTAL_FONT = Font(name=FONT_NAME, bold=True)
LINK_FONT = Font(name=FONT_NAME, color='008000')  # green = cross-sheet link
NORMAL_FONT = Font(name=FONT_NAME)


# ── Transaction ───────────────────────────────────────────────────────────────

SUPPORTED_ACTIONS = {"BUY", "SELL", "ROC", "SPLIT"}
ACTION_ALIASES = {
    "DIY_BUY": "BUY", "DIY BUY": "BUY",
    "DIY_SELL": "SELL", "DIY SELL": "SELL",
    "RETURNOFCAPITAL": "ROC", "RETURN OF CAPITAL": "ROC", "RETURN_OF_CAPITAL": "ROC",
    "CONSOLIDATION": "SPLIT", "REVERSE SPLIT": "SPLIT", "REVERSE_SPLIT": "SPLIT",
    "STOCK SPLIT": "SPLIT", "SPLIT/CONSOLIDATION": "SPLIT", "COMBINATION": "SPLIT",
    "REVERSE SPLIT/COMBINATION": "SPLIT",
}


class Transaction:
    def __init__(self, txn_date, ticker, action, qty, price, commission, currency, note=""):
        self.date = txn_date
        self.ticker = ticker.strip().upper()
        self.action = action.strip().upper()
        self.qty = qty
        self.price = price
        self.commission = commission
        self.currency = (currency or "CAD").strip().upper()
        self.note = note
        # filled in by compute_gains
        self.exchange_rate = Decimal('1')
        self.share_balance = Decimal('0')
        self.proceeds = Decimal('0')
        self.capital_gain = Decimal('0')
        self.acb = Decimal('0')


class Transactions:
    def __init__(self, iterable=None):
        self._txns = list(iterable or [])

    @property
    def tickers(self):
        return sorted({t.ticker for t in self._txns})

    def __len__(self): return len(self._txns)
    def __iter__(self): return iter(self._txns)
    def __bool__(self): return bool(self._txns)

    def filter_by(self, tickers=None):
        def keep(t):
            return not (tickers and t.ticker not in tickers)
        return Transactions(filter(keep, self._txns))


# ── Reading the input workbook ────────────────────────────────────────────────

def _to_date(value, context):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip().split()[0], '%Y-%m-%d').date()
    except (ValueError, IndexError):
        sys.exit(f"{context}: bad date '{value}' (expected YYYY-MM-DD or an Excel date cell)")


def _to_decimal(value, context, label):
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        sys.exit(f"{context}: invalid {label} '{value}'")


def _to_decimal_optional(value, context, label, default):
    """Blank/missing -> default. Present-but-garbage -> still a hard error
    (so a real typo doesn't get silently swallowed as 0)."""
    if value is None or str(value).strip() == "":
        return default
    return _to_decimal(value, context, label)


# Logical field -> accepted header name. settlementDate is checked before Date
# because when both exist (typical brokerage exports), settlement date is the
# one that actually matters for ACB/tax purposes, not the trade date.
COLUMN_ALIASES = {
    'date': ['settlementdate', 'settlement date', 'date'],
    'ticker': ['symbol', 'ticker', 'security'],
    'action': ['type', 'action', 'transaction type'],
    'quantity': ['quantity', 'qty', 'shares', 'units'],
    'price': ['price'],
    'currency': ['currency'],
    'commission': ['fee', 'commission'],      # optional; missing -> $0
}
REQUIRED_FIELDS = ['date', 'ticker', 'action', 'quantity', 'price', 'currency']


def _find_column(headers_lower, aliases):
    for alias in aliases:
        if alias in headers_lower:
            return headers_lower.index(alias)
    return None


def read_transactions_sheet(ws):
    """Returns (Transactions, skipped) where skipped is {raw_action: count} for any
    row whose action wasn't recognized — those rows are left out of the output
    entirely (not even shown in Summary), for things like
    dividends/transfers that have no ACB effect."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers_lower = [str(h).strip().lower() if h is not None else "" for h in header_row]

    cols = {field: _find_column(headers_lower, aliases) for field, aliases in COLUMN_ALIASES.items()}
    missing = [field for field in REQUIRED_FIELDS if cols[field] is None]
    if missing:
        sys.exit(f"Transactions sheet is missing required column(s): {missing}. "
                 f"Found headers: {list(header_row)}")
    if cols['commission'] is None:
        print("Note: no 'Fee' column found in Transactions sheet — "
              "assuming $0 commission/fee for every row.")

    transactions = []
    skipped = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[cols['date']] is None:            # skip blank trailing rows
            continue
        ctx = f"Transactions row {row_num}"
        raw_action = str(row[cols['action']] or "").strip().upper()
        action = ACTION_ALIASES.get(raw_action, raw_action)

        if action not in SUPPORTED_ACTIONS:
            skipped[raw_action] = skipped.get(raw_action, 0) + 1
            continue

        txn_date = _to_date(row[cols['date']], ctx)
        ticker = str(row[cols['ticker']] or "").strip()

        note = ""
        qty_raw = row[cols['quantity']]
        if action == 'ROC' and (qty_raw is None or str(qty_raw).strip() == ""):
            qty = Decimal('1')   # ROC: Price is the total amount, so a blank qty defaults to 1
        else:
            qty = _to_decimal(qty_raw, ctx, "quantity")
            if qty < 0:
                qty = abs(qty)   # source data may use negative qty for sells — Action is authoritative                

        price = _to_decimal(row[cols['price']], ctx, "price")
        if action == 'SPLIT':
            if price <= 0:
                sys.exit(f"{ctx}: SPLIT needs a positive 'old shares' count in Price "
                         f"(Quantity = new shares, Price = old shares — e.g. a 5-for-1 split "
                         f"is Quantity=5, Price=1).")
            if qty <= 0:
                sys.exit(f"{ctx}: SPLIT needs a positive 'new shares' count in Quantity.")
        elif price < 0:
            sys.exit(f"{ctx}: price cannot be negative ({price}). "
                     f"Use the Action column (BUY/SELL) for direction, not a negative price.")

        commission_raw = row[cols['commission']] if cols['commission'] is not None else None
        commission = abs(_to_decimal_optional(commission_raw, ctx, "commission", Decimal('0')))
        currency = str(row[cols['currency']] or "").strip().upper()
        if not currency:
            sys.exit(f"{ctx}: currency is required (e.g. CAD, USD)")

        transactions.append(Transaction(txn_date, ticker, action, qty, price, commission, currency, note))

    transactions.sort(key=lambda t: t.date)
    return Transactions(transactions), skipped


FX_COLUMN_RE = re.compile(r'^FX([A-Z]{3})CAD$')


def read_exchangerate_sheet(ws):
    """Accepts data pasted straight from a Bank of Canada Valet CSV export: a
    'date' column plus one or more FX<CCY>CAD columns (e.g. FXUSDCAD, FXEURCAD),
    one row per date."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers_upper = [str(h).strip().upper() if h is not None else "" for h in header_row]

    date_col = headers_upper.index('DATE') if 'DATE' in headers_upper else None
    if date_col is None:
        sys.exit("ExchangeRate sheet needs a 'date' column.")

    fx_cols = {}
    for i, h in enumerate(headers_upper):
        m = FX_COLUMN_RE.match(h)
        if m:
            fx_cols[m.group(1)] = i

    if not fx_cols:
        sys.exit("ExchangeRate sheet needs at least one FX<CCY>CAD column (e.g. "
                 "FXUSDCAD), pasted straight from the Bank of Canada Valet CSV export.")

    table = {}   # currency -> {date: Decimal}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[date_col] is None:
            continue
        ctx = f"ExchangeRate row {row_num}"
        d = _to_date(row[date_col], ctx)
        for currency, col in fx_cols.items():
            v = row[col]
            if v is not None and str(v).strip() != "":
                table.setdefault(currency, {})[d] = _to_decimal(v, ctx, f"{currency} rate")

    return table


def get_rate(rates_table, currency, d):
    if currency == 'CAD':
        return Decimal('1')
    table = rates_table.get(currency)
    if not table:
        sys.exit(f"No rates for currency '{currency}' in the ExchangeRate sheet.")
    if d in table:
        return table[d]
    preceding = [k for k in table if k < d]
    if preceding:
        return table[max(preceding)]
    sys.exit(f"No {currency} rate on or before {d}. "
             f"Earliest rate you provided is {min(table)} — add an earlier row.")


def read_workbook(path):
    wb = load_workbook(path, data_only=True)
    for required in ("Transactions", "ExchangeRate"):
        if required not in wb.sheetnames:
            sys.exit(f"Workbook is missing a '{required}' sheet. Found: {wb.sheetnames}")
    transactions, skipped = read_transactions_sheet(wb["Transactions"])
    rates = read_exchangerate_sheet(wb["ExchangeRate"])
    return transactions, rates, skipped


# ── ACB engine ────────────────────────────────────────────────────────────────

def _superficial_loss_ratio(txn, all_txns_list):
    """formula for partial dispositions:
        Superficial Loss = (min(S, P, B) / S) x Total Loss
    where S = shares sold in this disposition, P = total shares acquired during
    the 61-day window (30 days before to 30 days after), and B = shares of this
    ticker still held at the END of that window.

    A SPLIT changes the share denomination, so a raw share count from a different
    point in time isn't directly comparable to S. Every quantity is converted into
    "as of this sale" terms via the cumulative split ratio between that point and
    the sale, before S/P/B are compared — e.g. if you sell 100 shares, still hold
    900, and a 10-for-1 reverse split then leaves you with 90, B is 900 (the same
    denomination as the 100 sold), not the raw post-split 90.

    Returns (ratio, S, P, B) — ratio is 0 if no part of the loss is superficial.
    """
    S = txn.qty
    if txn.capital_gain >= 0:
        return Decimal('0'), S, Decimal('0'), Decimal('0')

    lo, hi = txn.date - timedelta(days=30), txn.date + timedelta(days=30)
    window = [t for t in all_txns_list if lo <= t.date <= hi]
    idx = window.index(txn)

    # cum_factor[i] = cumulative split ratio from the start of the window up to
    # (but not including) position i's own split, if i itself is a SPLIT.
    cum_factor = []
    factor = Decimal('1')
    for t in window:
        cum_factor.append(factor)
        if t.action == 'SPLIT':
            factor *= (t.qty / t.price)
    sale_factor = cum_factor[idx]

    # P: every BUY in the window, rescaled into as-of-the-sale terms.
    P = Decimal('0')
    for i, t in enumerate(window):
        if t.action == 'BUY':
            P += t.qty * (sale_factor / cum_factor[i])
    if P == 0:
        return Decimal('0'), S, P, Decimal('0')

    # B: walk the ACTUAL share balance forward to the end of the window (so SPLITs
    # apply correctly in real time), then rescale the result back into as-of-the-sale
    # terms using the cumulative split ratio between the sale and the window's end.
    B_actual = txn.share_balance
    end_factor = sale_factor
    for t in window[idx + 1:]:
        if t.action == 'BUY':
            B_actual += t.qty
        elif t.action == 'SELL':
            B_actual -= t.qty
        elif t.action == 'SPLIT':
            ratio = t.qty / t.price
            B_actual *= ratio
            end_factor *= ratio
        # ROC doesn't change share count — no-op
    B = B_actual * (sale_factor / end_factor)
    if B <= 0:
        return Decimal('0'), S, P, B

    smallest = min(S, P, B)
    return smallest / S, S, P, B


def _fmt_qty(d):
    return f"{d.normalize():f}" if d == d.to_integral() else f"{d:f}"


def compute_gains(transactions, rates_table, ledger):
    """Runs the ACB engine for one ticker's transactions (chronological order)
    and appends audit-trail dicts into `ledger`. Each real transaction gets its
    own row. When a (partial) superficial loss applies, a separate synthetic
    "adjustment" row is appended immediately below it — right after the sell
    if shares remain afterward, or below the first repurchase after a full
    disposition (since there's nothing to attach the ACB to until then)."""
    txn_list = list(transactions)
    share_balance = Decimal('0')
    total_acb = Decimal('0')
    pending = None        # holds a queued adjustment waiting for the next BUY
    acq_year_min = None   # tracks which calendar year(s) of BUYs contribute to the
    acq_year_max = None   # currently-held pool, for Schedule 3's "Year acquired" column

    for t in txn_list:
        rate = get_rate(rates_table, t.currency, t.date)
        t.exchange_rate = rate
        total_acb_before = total_acb
        acb_per_share_before = _r4(total_acb / share_balance) if share_balance else Decimal('0')
        proceeds = _r2(t.qty * t.price * rate)
        commission_cad = _r2(t.commission * rate)
        year_acquired = None

        if t.action == 'SELL':
            if t.qty > share_balance:
                sys.exit(f"{t.date} {t.ticker}: selling {t.qty} shares but only "
                         f"{share_balance} on hand — check for a missing BUY or wrong order.")
            share_balance -= t.qty
            if share_balance == 0:
                # Full disposition — use the exact remaining ACB pool rather than
                # acb_per_share_before * qty, which can leave a stray cent or two
                # behind once acb_per_share_before has been rounded to 4dp.
                acb = total_acb_before
            else:
                acb = _r2(acb_per_share_before * t.qty)
            capital_gain = _r2(proceeds - commission_cad - acb)
            total_acb = _r2(total_acb - acb)
            year_acquired = (str(acq_year_min) if acq_year_min == acq_year_max
                              else f"Various ({acq_year_min}-{acq_year_max})")
        elif t.action == 'BUY':
            share_balance += t.qty
            acb = _r2(proceeds + commission_cad)
            capital_gain = Decimal('0')
            total_acb = _r2(total_acb + acb)
            acq_year_min = t.date.year if acq_year_min is None else min(acq_year_min, t.date.year)
            acq_year_max = t.date.year if acq_year_max is None else max(acq_year_max, t.date.year)
        elif t.action == 'ROC':
            # Return of capital reduces ACB without changing share count. If it
            # would push ACB negative, the excess is a deemed capital gain
            # (CRA s.53(2)(h)) and ACB resets to $0 — it is NOT a disposition,
            # so it never feeds the superficial-loss S/P/B calculation.
            acb = _r2(-proceeds)
            total_acb = _r2(total_acb - proceeds)
            capital_gain = Decimal('0')
            if total_acb < 0:
                capital_gain = _r2(-total_acb)
                total_acb = Decimal('0')
        elif t.action == 'SPLIT':
            ratio = t.qty / t.price
            old_balance = share_balance
            share_balance *= ratio   # total ACB is unchanged; ACB/share adjusts via the divide
            proceeds = Decimal('0')
            acb = Decimal('0')
            capital_gain = Decimal('0')
        else:
            sys.exit(f"Internal error: unhandled action '{t.action}' for {t.ticker} on {t.date}.")

        t.share_balance = share_balance
        t.proceeds = proceeds
        t.capital_gain = capital_gain
        t.acb = acb

        note = t.note
        highlight = False
        ratio_out = None
        full_gain_loss = None

        if t.action == 'SELL':
            full_gain_loss = capital_gain   # raw, pre-adjustment — the writer derives this via
            ratio_out = Decimal('0')        # formula instead, but keep it here too for clarity
            if capital_gain < 0:
                ratio_out, S, P, B = _superficial_loss_ratio(t, txn_list)
                if ratio_out > 0:
                    total_loss = -capital_gain
                    superficial_amount = _r2(ratio_out * total_loss)
                    new_capital_gain = _r2(capital_gain + superficial_amount)   # less negative
                    t.capital_gain = new_capital_gain
                    capital_gain = new_capital_gain
                    highlight = True
                    pct = ratio_out * 100
                    if share_balance > 0:
                        note = (note + "; " if note else "") + (
                            f"superficial loss rule (partial): {pct:.1f}% of the {total_loss:,.2f} "
                            f"loss is denied — see adjustment row below"
                        )
                    else:
                        note = (note + "; " if note else "") + (
                            f"superficial loss rule (partial): {pct:.1f}% of the {total_loss:,.2f} "
                            f"loss is denied — adjustment carried to the first repurchase after this sale"
                        )
        elif t.action == 'ROC' and capital_gain > 0:
            full_gain_loss = capital_gain   # the deemed gain itself
            note = (note + "; " if note else "") + (
                f"return of capital exceeded remaining ACB — {capital_gain:,.2f} deemed capital "
                f"gain realized (CRA s.53(2)(h)), ACB reset to $0.00"
            )
        elif t.action == 'SPLIT':
            note = (note + "; " if note else "") + (
                f"{_fmt_qty(t.qty)}-for-{_fmt_qty(t.price)} {'split' if ratio >= 1 else 'reverse split'} "
                f"applied: {_fmt_qty(old_balance)} -> {_fmt_qty(share_balance)} shares "
                f"(total ACB unchanged, ACB/share adjusts inversely)"
            )

        ledger.append({
            'ticker': t.ticker, 'date': t.date, 'action': t.action, 'qty': t.qty, 'price': t.price,
            'currency': t.currency, 'rate': rate, 'proceeds_cad': proceeds, 'commission_cad': commission_cad,
            'acb_per_share_before': acb_per_share_before,
            'acb_change': _r2(total_acb - total_acb_before),
            'total_acb_after': total_acb, 'share_balance_after': share_balance,
            'capital_gain': capital_gain, 'highlight': highlight, 'kind': 'transaction',
            'ratio': ratio_out, 'full_gain_loss': full_gain_loss, 'note': note,
            'year_acquired': year_acquired,
        })
        source_idx = len(ledger) - 1   # index of the row just appended, for adjustment cross-references

        if t.action == 'SELL' and highlight:
            if share_balance > 0:
                # Shares remain — attach the ACB adjustment right here, immediately.
                acb_per_share_pre_adj = _r4(total_acb / share_balance) if share_balance else Decimal('0')
                total_acb = _r2(total_acb + superficial_amount)
                ledger.append({
                    'ticker': t.ticker, 'date': t.date, 'action': 'ADJ',
                    'description': f"↳ S/L adjustment ({t.date} sale)",
                    'qty': None, 'price': None, 'currency': None, 'rate': None,
                    'proceeds_cad': None, 'commission_cad': None,
                    'acb_per_share_before': acb_per_share_pre_adj,
                    'acb_change': superficial_amount, 'total_acb_after': total_acb,
                    'share_balance_after': share_balance, 'capital_gain': None,
                    'highlight': True, 'kind': 'adjustment', 'ratio': None,
                    'full_gain_loss': None, 'source_idx': source_idx, 'year_acquired': None,
                    'note': (f"{superficial_amount:,.2f} of the {total_loss:,.2f} loss on the "
                             f"{t.date} sale of {_fmt_qty(S)} shares added back to ACB of the "
                             f"{_fmt_qty(share_balance)} remaining shares — "
                             f"min(S={_fmt_qty(S)}, P={_fmt_qty(P)}, B={_fmt_qty(B)})/S = "
                             f"{ratio_out * 100:.1f}% per the superficial loss rule for partial dispositions."),
                })
            else:
                # Full disposition — nothing to attach to yet; queue for the next BUY.
                pending = {'amount': superficial_amount, 'sale_date': t.date,
                           'S': S, 'P': P, 'B': B, 'ratio': ratio_out, 'total_loss': total_loss,
                           'source_idx': source_idx}

        elif t.action == 'BUY' and pending is not None:
            acb_per_share_pre_adj = _r4(total_acb / share_balance) if share_balance else Decimal('0')
            total_acb = _r2(total_acb + pending['amount'])
            ledger.append({
                'ticker': t.ticker, 'date': t.date, 'action': 'ADJ',
                'description': f"↳ S/L adjustment ({pending['sale_date']} sale)",
                'qty': None, 'price': None, 'currency': None, 'rate': None,
                'proceeds_cad': None, 'commission_cad': None,
                'acb_per_share_before': acb_per_share_pre_adj,
                'acb_change': pending['amount'], 'total_acb_after': total_acb,
                'share_balance_after': share_balance, 'capital_gain': None,
                'highlight': True, 'kind': 'adjustment', 'ratio': None,
                'full_gain_loss': None, 'source_idx': pending['source_idx'], 'year_acquired': None,
                'note': (f"{pending['amount']:,.2f} of the {pending['total_loss']:,.2f} loss on the "
                         f"{pending['sale_date']} sale (full disposition at the time) added back to "
                         f"ACB of these {_fmt_qty(t.qty)} newly purchased shares — "
                         f"min(S={_fmt_qty(pending['S'])}, P={_fmt_qty(pending['P'])}, "
                         f"B={_fmt_qty(pending['B'])})/S = {pending['ratio'] * 100:.1f}% per the superficial "
                         f"loss rule for partial dispositions."),
            })
            pending = None

    if pending is not None:
        sys.exit(f"Internal error: unresolved superficial-loss adjustment for {pending['sale_date']} "
                 f"— a qualifying repurchase was detected but never reached. Please report this.")


# ── Excel output: Summary sheet ───────────────────────────────────────────────

SUMMARY_HEADERS = [
    "Date", "Action", "Qty", "Price", "Currency", "FX Rate",
    "Proceeds/Cost (CAD)", "Commission (CAD)", "ACB/Share Before",
    "ACB Change", "Total ACB After", "Share Bal. After", "ACB/Share After",
    "Full G/L", "SL %", "SL Amount", "Net G/L", "Reportable", "Notes",
]
(COL_DATE, COL_ACTION, COL_QTY, COL_PRICE, COL_CURRENCY, COL_RATE,
 COL_PROCEEDS, COL_COMMISSION, COL_ACB_PS_BEFORE, COL_ACB_CHANGE,
 COL_TOTAL_ACB_AFTER, COL_SHARE_BAL_AFTER, COL_ACB_PS_AFTER, COL_FULL_GL,
 COL_SL_PCT, COL_SL_AMOUNT, COL_NET_GL, COL_REPORTABLE, COL_NOTES) = range(1, len(SUMMARY_HEADERS) + 1)
NUM_SUMMARY_COLS = len(SUMMARY_HEADERS)

MONEY_COLS = {COL_PROCEEDS, COL_COMMISSION, COL_ACB_CHANGE, COL_TOTAL_ACB_AFTER,
              COL_FULL_GL, COL_SL_AMOUNT, COL_NET_GL}
FOUR_DP_COLS = {COL_ACB_PS_BEFORE, COL_ACB_PS_AFTER}


def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 42


def _col_width(c):
    return {COL_DATE: 11, COL_ACTION: 8, COL_QTY: 9, COL_PRICE: 8, COL_CURRENCY: 8,
            COL_RATE: 8, COL_PROCEEDS: 11, COL_COMMISSION: 10, COL_ACB_PS_BEFORE: 11,
            COL_ACB_CHANGE: 11, COL_TOTAL_ACB_AFTER: 11, COL_SHARE_BAL_AFTER: 10,
            COL_ACB_PS_AFTER: 11, COL_FULL_GL: 11, COL_SL_PCT: 7, COL_SL_AMOUNT: 9,
            COL_NET_GL: 11, COL_REPORTABLE: 10, COL_NOTES: 42}.get(c, 10)


def _write_summary_sheet(wb, all_ledgers):
    """Returns reportable_index: {year: [(ticker, ledger_row_dict, summary_row_number), ...]}"""
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A1"
    for c in range(1, NUM_SUMMARY_COLS + 1):
        ws.column_dimensions[L(c)].width = _col_width(c)
    ws.column_dimensions[L(COL_ACB_PS_BEFORE)].hidden = True   # audit trail only, not used by any formula

    row_idx = 1
    reportable_index = {}   # year -> list of (ticker, row_dict, summary_row)

    for ticker, ledger in all_ledgers:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=NUM_SUMMARY_COLS)
        band = ws.cell(row=row_idx, column=1, value=f"Ticker: {ticker}")
        band.font = BAND_FONT
        band.fill = BAND_FILL
        row_idx += 1

        for c, h in enumerate(SUMMARY_HEADERS, start=1):
            ws.cell(row=row_idx, column=c, value=h)
        _style_header_row(ws, row_idx, NUM_SUMMARY_COLS)
        row_idx += 1
        first_detail_row = row_idx
        years_seen = set()

        for r in ledger:
            is_reportable_event = (
                r['kind'] == 'transaction' and
                (r['action'] == 'SELL' or (r['action'] == 'ROC' and r['full_gain_loss'] is not None))
            )
            plain_values = {
                COL_DATE: r['date'], COL_ACTION: r.get('description') or r['action'],
                COL_QTY: r['qty'], COL_PRICE: r['price'], COL_CURRENCY: r['currency'],
                COL_RATE: r['rate'], COL_PROCEEDS: r['proceeds_cad'], COL_COMMISSION: r['commission_cad'],
                COL_ACB_PS_BEFORE: r['acb_per_share_before'], COL_ACB_CHANGE: r['acb_change'],
                COL_TOTAL_ACB_AFTER: r['total_acb_after'], COL_SHARE_BAL_AFTER: r['share_balance_after'],
                COL_NOTES: r['note'],
            }
            if is_reportable_event:
                plain_values[COL_REPORTABLE] = 'Y'
            for c, v in plain_values.items():
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.font = NORMAL_FONT
                if c == COL_DATE: cell.number_format = DATE_FMT
                elif c == COL_QTY: cell.number_format = QTY_FMT
                elif c == COL_RATE: cell.number_format = RATE_FMT
                elif c in FOUR_DP_COLS: cell.number_format = MONEY_FMT_4DP
                elif c in MONEY_COLS: cell.number_format = MONEY_FMT
                if r['highlight']:
                    cell.fill = SL_FILL

            formula_cells = {
                COL_ACB_PS_AFTER: f'=ROUND(IFERROR({L(COL_TOTAL_ACB_AFTER)}{row_idx}/'
                                   f'{L(COL_SHARE_BAL_AFTER)}{row_idx},0),4)',
            }
            if r['kind'] == 'transaction' and r['action'] == 'SELL':
                formula_cells[COL_FULL_GL] = (
                    f'=ROUND({L(COL_PROCEEDS)}{row_idx}-{L(COL_COMMISSION)}{row_idx}+'
                    f'{L(COL_ACB_CHANGE)}{row_idx},2)')
                if r['full_gain_loss'] is not None and r['full_gain_loss'] < 0:
                    formula_cells[COL_SL_PCT] = r['ratio']
                    formula_cells[COL_SL_AMOUNT] = (
                        f'=ROUND(-{L(COL_FULL_GL)}{row_idx}*{L(COL_SL_PCT)}{row_idx},2)')
                formula_cells[COL_NET_GL] = (
                    f'=ROUND({L(COL_FULL_GL)}{row_idx}+{L(COL_SL_AMOUNT)}{row_idx},2)')
            elif r['kind'] == 'adjustment':
                src_row = first_detail_row + r['source_idx']
                formula_cells[COL_FULL_GL] = f'={L(COL_FULL_GL)}{src_row}'
                formula_cells[COL_SL_PCT] = f'={L(COL_SL_PCT)}{src_row}'
                formula_cells[COL_SL_AMOUNT] = f'={L(COL_SL_AMOUNT)}{src_row}'
            elif r['full_gain_loss'] is not None:        # ROC deemed gain
                formula_cells[COL_FULL_GL] = r['full_gain_loss']
                formula_cells[COL_NET_GL] = r['full_gain_loss']
            else:
                formula_cells[COL_NET_GL] = Decimal('0')

            for c, v in formula_cells.items():
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.font = NORMAL_FONT
                cell.number_format = PCT_FMT if c == COL_SL_PCT else MONEY_FMT
                if r['highlight']:
                    cell.fill = SL_FILL

            if is_reportable_event:
                yr = r['date'].year
                years_seen.add(yr)
                reportable_index.setdefault(yr, []).append((ticker, r, row_idx))

            row_idx += 1

        last_detail_row = row_idx - 1
        for yr in sorted(years_seen):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            ws.cell(row=row_idx, column=1, value=f"Reportable {yr} gain/(loss) for {ticker}:").font = SUBTOTAL_FONT
            cell = ws.cell(
                row=row_idx, column=COL_NET_GL,
                value=(f'=ROUND(SUMPRODUCT(({L(COL_REPORTABLE)}{first_detail_row}:'
                       f'{L(COL_REPORTABLE)}{last_detail_row}="Y")*'
                       f'(YEAR({L(COL_DATE)}{first_detail_row}:{L(COL_DATE)}{last_detail_row})={yr})*'
                       f'{L(COL_NET_GL)}{first_detail_row}:{L(COL_NET_GL)}{last_detail_row}),2)')
            )
            cell.font = SUBTOTAL_FONT
            cell.number_format = MONEY_FMT
            row_idx += 1
        row_idx += 1   # blank spacer row before next ticker

    return reportable_index


# ── Excel output: Schedule 3 sheets ──────────────────────────────────────────

SCHED3_HEADERS = [
    "Date", "Number", "Name of fund/corp. and class of shares", "(1) Year acquired",
    "(2) Proceeds of disposition", "(3) Adjusted cost base", "(4) Outlays and expenses",
    "(5) Gain or loss", "Notes",
]


def _write_schedule3_sheet(wb, year, entries):
    """entries: list of (ticker, ledger_row_dict, summary_row_number) for this year,
    pulled from _write_summary_sheet's reportable_index. Cross-references the Summary
    sheet by formula for columns 2-5 so the figures stay tied to one source of truth."""
    ws = wb.create_sheet(f"Schedule 3 - {year}", 0)
    for c, h in enumerate(SCHED3_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(SCHED3_HEADERS))
    widths = [11, 9, 14, 13, 16, 16, 16, 13, 40]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[L(c)].width = w

    rows = sorted(entries, key=lambda e: (e[1]['date'], e[0]))
    for i, (ticker, r, summary_row) in enumerate(rows, start=2):
        is_roc = (r['action'] == 'ROC')
        name = f"{ticker} (ROC deemed gain)" if is_roc else ticker
        vals = [
            r['date'],
            None if is_roc else r['qty'],
            name,
            None if is_roc else r['year_acquired'],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = NORMAL_FONT
            if c == 1: cell.number_format = DATE_FMT
            elif c == 2: cell.number_format = QTY_FMT

        if not is_roc:
            proceeds_cell = ws.cell(row=i, column=5, value=f"=Summary!{L(COL_PROCEEDS)}{summary_row}")
            acb_cell = ws.cell(row=i, column=6, value=f"=-Summary!{L(COL_ACB_CHANGE)}{summary_row}")
            outlay_cell = ws.cell(row=i, column=7, value=f"=Summary!{L(COL_COMMISSION)}{summary_row}")
            for cell in (proceeds_cell, acb_cell, outlay_cell):
                cell.font = LINK_FONT
                cell.number_format = MONEY_FMT
        gain_cell = ws.cell(row=i, column=8, value=f"=Summary!{L(COL_NET_GL)}{summary_row}")
        gain_cell.font = LINK_FONT
        gain_cell.number_format = MONEY_FMT

        note = ""
        if r['highlight']:
            note = "Superficial loss rule applied — see Summary tab for the S/P/B calculation."
        elif is_roc:
            note = "Deemed capital gain from return of capital exceeding ACB (CRA s.53(2)(h))."
        note_cell = ws.cell(row=i, column=9, value=note)
        note_cell.font = NORMAL_FONT
        if r['highlight']:
            for c in range(1, 10):
                ws.cell(row=i, column=c).fill = SL_FILL

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=7, value="Total:").font = SUBTOTAL_FONT
    total_cell = ws.cell(row=total_row, column=8, value=f"=ROUND(SUM(H2:H{total_row - 1}),2)")
    total_cell.font = SUBTOTAL_FONT
    total_cell.number_format = MONEY_FMT


# ── Excel output: T1135 sheets (one per year) ────────────────────────────────

MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def _month_range(start_ym, end_ym):
    y, m = start_ym
    out = []
    while (y, m) <= end_ym:
        out.append((y, m))
        m += 1
        if m == 13:
            m, y = 1, y + 1
    return out


def _monthly_peak_series(snapshots, months):
    """snapshots: chronological list of (date, total_acb_after) — already rounded
    to the cent by compute_gains. Returns {(y,m): peak}."""
    result = {}
    carry = Decimal('0')
    idx = 0
    for (y, m) in months:
        month_end = (date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)) - timedelta(days=1)
        peak = carry
        while idx < len(snapshots) and snapshots[idx][0] <= month_end:
            peak = max(peak, snapshots[idx][1])
            carry = snapshots[idx][1]
            idx += 1
        result[(y, m)] = peak
    return result


def _write_t1135_sheets(wb, all_ledgers):
    """One 'T1135 - <year>' tab per calendar year touched by any holding. Rows are
    tickers — ALL of them, including CAD ones, since currency alone doesn't decide
    specified-foreign-property status (e.g. a CAD-hedged CDR of a foreign company
    is still reportable despite trading in CAD). CAD tickers default to NOT
    reportable, non-CAD default to reportable; change via T1135 Reportable Column."""
    tickers_info = []   # (ticker, snapshots, currency_label, reportable)
    for ticker, ledger in all_ledgers:
        txn_rows = [r for r in ledger if r['kind'] == 'transaction' and r['action'] in ('BUY', 'SELL', 'ROC')]
        if not txn_rows:
            continue
        currencies = sorted({r['currency'] for r in txn_rows if r['currency']})
        currency_label = '/'.join(currencies) if currencies else 'CAD'
        reportable = any(c != 'CAD' for c in currencies)
        snapshots = [(r['date'], r['total_acb_after']) for r in ledger if r['total_acb_after'] is not None]
        tickers_info.append((ticker, snapshots, currency_label, reportable))

    if not tickers_info:
        return []

    all_dates = [s[0] for _, snaps, _, _ in tickers_info for s in snaps]
    start_ym = (min(all_dates).year, min(all_dates).month)
    # A holding carries forward unchanged once transactions stop, so the range
    # must reach at least today's month — not just the month of the last
    # recorded transaction — or a quiet ticker would silently drop out of the
    # current (still-in-progress) year's monthly columns, and an entire year
    # with zero activity but a still-held position would be skipped outright.
    # We never project past today: future months are genuinely unknown, not
    # "assumed unchanged," so they're left out rather than guessed at.
    today = date.today()
    end_ym = max((max(all_dates).year, max(all_dates).month), (today.year, today.month))
    months = _month_range(start_ym, end_ym)
    series = {ticker: _monthly_peak_series(snaps, months) for ticker, snaps, _, _ in tickers_info}

    years = sorted({y for y, m in months})
    written = []
    for year in years:
        year_months = [(y, m) for (y, m) in months if y == year]
        rows = sorted(
            (t, cur, rep) for t, snaps, cur, rep in tickers_info
            if any(series[t][(y, m)] != 0 for (y, m) in year_months)
        )
        if not rows:
            continue
        _write_one_t1135_year(wb, year, year_months, series, rows)
        written.append(year)

    return written


def _write_one_t1135_year(wb, year, year_months, series, rows):
    ws = wb.create_sheet(f"T1135 - {year}", 0)
    COL_TICKER, COL_CCY, COL_REPORT = 1, 2, 3
    n_months = len(year_months)
    first_month_col = 4
    last_month_col = 3 + n_months
    peak_col = 4 + n_months

    headers = (["Ticker", "Currency", "T1135 Reportable?"] +
               [MONTH_NAMES[m - 1] for (_, m) in year_months] + ["Peak for Year"])
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.column_dimensions[L(COL_TICKER)].width = 12
    ws.column_dimensions[L(COL_CCY)].width = 10
    ws.column_dimensions[L(COL_REPORT)].width = 11
    for c in range(first_month_col, last_month_col + 1):
        ws.column_dimensions[L(c)].width = 11
    ws.column_dimensions[L(peak_col)].width = 12

    first_data_row = 2
    row_idx = first_data_row
    for ticker, currency_label, reportable in rows:
        ws.cell(row=row_idx, column=COL_TICKER, value=ticker).font = NORMAL_FONT
        ws.cell(row=row_idx, column=COL_CCY, value=currency_label).font = NORMAL_FONT
        ws.cell(row=row_idx, column=COL_REPORT, value='Yes' if reportable else 'No').font = NORMAL_FONT
        for i, (y, m) in enumerate(year_months):
            v = series[ticker][(y, m)]   # already rounded to the cent
            cell = ws.cell(row=row_idx, column=first_month_col + i, value=v)
            cell.font = NORMAL_FONT
            cell.number_format = MONEY_FMT
        peak_cell = ws.cell(row=row_idx, column=peak_col, value=(
            f"=ROUND(MAX({L(first_month_col)}{row_idx}:{L(last_month_col)}{row_idx}),2)"))
        peak_cell.font = NORMAL_FONT
        peak_cell.number_format = MONEY_FMT
        row_idx += 1
    last_data_row = row_idx - 1

    # Gray out excluded (Reportable = No) rows — a live conditional format, not a static
    # fill, so flipping Yes/No directly in Excel updates the look too.
    full_range = f"A{first_data_row}:{L(peak_col)}{last_data_row}"
    ws.conditional_formatting.add(
        full_range, FormulaRule(formula=[f'${L(COL_REPORT)}{first_data_row}="No"'], fill=EXCLUDED_FILL))

    row_idx += 1
    total_row = row_idx
    ws.cell(row=total_row, column=1, value="Total (Reportable tickers)").font = SUBTOTAL_FONT
    for i in range(n_months):
        col = first_month_col + i
        cl = L(col)
        formula = (f'=ROUND(SUMIFS({cl}{first_data_row}:{cl}{last_data_row},'
                   f'{L(COL_REPORT)}{first_data_row}:{L(COL_REPORT)}{last_data_row},"Yes"),2)')
        cell = ws.cell(row=total_row, column=col, value=formula)
        cell.font = SUBTOTAL_FONT
        cell.number_format = MONEY_FMT
    peak_total_cell = ws.cell(row=total_row, column=peak_col, value=(
        f"=ROUND(MAX({L(first_month_col)}{total_row}:{L(last_month_col)}{total_row}),2)"))
    peak_total_cell.font = SUBTOTAL_FONT
    peak_total_cell.number_format = MONEY_FMT

    # Highlight any month (or the year's peak) that exceeds the T1135 threshold —
    # also a live conditional format, so it stays correct after Yes/No edits.
    total_range = f"{L(first_month_col)}{total_row}:{L(peak_col)}{total_row}"
    ws.conditional_formatting.add(
        total_range, CellIsRule(operator='greaterThan', formula=[str(T1135_THRESHOLD)], fill=WARN_FILL))


# ── Top-level orchestration ───────────────────────────────────────────────────

def write_output_workbook(input_path, output_path, all_ledgers):
    wb = load_workbook(input_path)
    reportable_index = _write_summary_sheet(wb, all_ledgers)
    t1135_years = _write_t1135_sheets(wb, all_ledgers)
    for year in sorted(reportable_index):
        _write_schedule3_sheet(wb, year, reportable_index[year])
    wb.save(output_path)
    return sorted(reportable_index), t1135_years


def run(xlsx_file):
    transactions, rates_table, skipped = read_workbook(xlsx_file)
    if not transactions:
        sys.exit("No matching transactions found.")

    all_ledgers = []
    for ticker in transactions.tickers:
        ticker_txns = transactions.filter_by(tickers=[ticker])
        ledger = []
        compute_gains(ticker_txns, rates_table, ledger)
        all_ledgers.append((ticker, ledger))

    out_path = str(Path(xlsx_file).with_name(f"{Path(xlsx_file).stem}_taxReview.xlsx"))
    years, t1135_years = write_output_workbook(xlsx_file, out_path, all_ledgers)
    msg = f"Wrote {out_path}"
    msg += f" — Schedule 3 years: {years}" if years else " — no reportable dispositions found"
    msg += f" — T1135 years: {t1135_years}" if t1135_years else ""
    print(msg)
    if skipped:
        details = ', '.join(f"{action!r}: {n}" for action, n in sorted(skipped.items()))
        print(f"Skipped {sum(skipped.values())} row(s) with unrecognized actions (not in output): {details}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Canadian Capital Gains / ACB Calculator (Excel-based)")
    parser.add_argument('xlsx_file', metavar='XLSX')
    args = parser.parse_args()
    run(args.xlsx_file)


if __name__ == '__main__':
    main()
